"""openpyxl isolation layer (plan section 14, bottom of the boundary).

The only module that imports openpyxl. Loading keeps formulas as
formulas (data_only=False default), so untouched cells survive
save/reopen exactly.
"""

import re

from openpyxl import load_workbook

# A1-style single-cell reference: 1-3 column letters + positive row.
_CELL_REF = re.compile(r"^([A-Za-z]{1,3})([1-9]\d*)$")

MAX_ROW = 1_048_576
MAX_COLUMN = "XFD"


def normalize_cell(cell):
    match = _CELL_REF.fullmatch(cell or "")
    if match is None:
        return None
    column, row = match.group(1).upper(), int(match.group(2))
    if row > MAX_ROW:
        return None
    if len(column) == len(MAX_COLUMN) and column > MAX_COLUMN:
        return None
    return f"{column}{row}"


def column_of(cell_ref):
    return _CELL_REF.fullmatch(cell_ref).group(1).upper()


def row_of(cell_ref):
    return int(_CELL_REF.fullmatch(cell_ref).group(2))


def open_draft(path):
    return load_workbook(path)


def save_draft(workbook, path):
    workbook.save(path)


def open_template(path):
    # Read side of open_draft: the template is never saved back.
    return load_workbook(path)


def has_sheet(workbook, name):
    return name in workbook.sheetnames


def sheet_names(workbook):
    return list(workbook.sheetnames)


def outline_rows(workbook, sheet):
    """Column letter and text of every non-empty cell in the used rows."""
    rows = {}
    # Worksheet.iter_rows expands the full max_row × max_column rectangle.
    # A distant sparse cell can make that rectangle enormous, so this
    # openpyxl isolation layer reads its sparse cell store directly.
    for (row, column), cell in workbook[sheet]._cells.items():
        if cell.value is None:
            continue
        value = str(cell.value).strip()
        if not value:
            continue
        rows.setdefault(row, []).append((column, cell.column_letter, value))
    return [
        (
            row,
            [(column_letter, value) for _, column_letter, value in sorted(cells)],
        )
        for row, cells in sorted(rows.items())
    ]


def read_cell(workbook, sheet, cell_ref):
    return workbook[sheet][cell_ref].value


def max_row(workbook, sheet):
    return workbook[sheet].max_row


def write_cell(workbook, sheet, cell_ref, value):
    workbook[sheet][cell_ref] = value
