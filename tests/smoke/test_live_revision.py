"""Live Revision and full-live-pipeline smoke tests (ticket #12).

These spend real Claude and Codex subscription quota and are excluded
from the default run; execute them with `pytest -m smoke`.

The revision-slice test controls the Revision inputs completely (fake
filler and reviewer) so the live decisions are checkable: a WARN
without a recommended value can only be answered with REBUT (the
verdict's action table), which deterministically exercises the single
live re-review round; a FAIL over genuinely conflicting sources sets
up CLEAR + note_append. The full-run test then drives every role live
through pause and resume on the clean sample workspace.
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from workflow_app.cli import build_runtimes
from workflow_app.models.review import ReReviewResult
from workflow_app.models.revision import RevisionResult
from workflow_app.runtimes.claude_code import (
    PROMPTS_DIR,
    ClaudeCodeRuntime,
)
from workflow_app.runtimes.claude_code import (
    ROLES as CLAUDE_ROLES,
)
from workflow_app.runtimes.codex import CodexRuntime
from workflow_app.runtimes.fake import FakeAgentRuntime
from workflow_app.workflow.engine import resume_workflow, run_workflow
from workflow_app.workspace import RunInputs, Workspace

pytestmark = pytest.mark.smoke

SHEET = "7) Practicum Courses"
BRIEF_PATH = "India 2008/Project_Brief.txt"
NOTE_PATH = "India 2008/Annual_Report_Note.txt"

CONFLICTING_NOTE = (
    "Annual report note: program operations commenced in March 2009,\n"
    "after a year of preparatory planning.\n"
)


def evidence(source_file, text):
    return {
        "source_file": source_file,
        "source_location": None,
        "evidence_text": text,
        "evidence_type": "direct",
    }


def proposal(column_name, cell, value, confidence, text):
    return {
        "sheet": SHEET,
        "row": 2,
        "column_name": column_name,
        "cell": cell,
        "value": value,
        "evidence": [evidence(BRIEF_PATH, text)],
        "rules_applied": ["extraction_rules"],
        "confidence": confidence,
        "status": "proposed",
        "notes": None,
    }


FILLER_FIXTURE = {
    "proposals": [
        proposal("Project ID*", "A2", "PRJ-2008", 0.80, "PRJ-<start year> rule"),
        proposal(
            "Start Date", "D2", "2008-03-15", 0.80, "began operations on 2008-03-15"
        ),
        proposal("Maturity", "E2", "Established", 0.70, "well established"),
        proposal(
            "Main Issue Area(s)", "G2", "Healthcare", 0.80, "community healthcare"
        ),
    ]
}

REVIEW_FIXTURE = {
    "findings": [
        {
            "cell": "A2",
            "verdict": "PASS",
            "issue_type": None,
            "current_value": "PRJ-2008",
            "recommended_value": None,
            "evidence": [evidence(BRIEF_PATH, "started 2008 -> PRJ-2008")],
            "reviewer_comment": "Matches the constructed-ID rule.",
            "missed_data": False,
        },
        {
            # The workspace's two sources genuinely conflict on this
            # value and neither dominates: the textbook CLEAR case.
            "cell": "D2",
            "verdict": "FAIL",
            "issue_type": "source_conflict",
            "current_value": "2008-03-15",
            "recommended_value": None,
            "evidence": [
                evidence(BRIEF_PATH, "began operations on 2008-03-15"),
                evidence(NOTE_PATH, "operations commenced in March 2009"),
            ],
            "reviewer_comment": (
                "The brief and the annual report note contradict each"
                " other on the start date (2008-03-15 vs March 2009);"
                " no single supported value is determinable."
            ),
            "missed_data": False,
        },
        {
            # No recommended value, so the action table leaves exactly
            # ACCEPT (illegal without a recommendation) or REBUT.
            "cell": "E2",
            "verdict": "WARN",
            "issue_type": "judgment_doubt",
            "current_value": "Established",
            "recommended_value": None,
            "evidence": [evidence(BRIEF_PATH, "first year of operations")],
            "reviewer_comment": (
                "One year of operations may be too short for"
                " 'Established' on the maturity scale; the brief's"
                " wording might refer to the partner network only."
            ),
            "missed_data": False,
        },
    ]
}


def run_inputs(inputs, scoping_answers=None):
    return RunInputs(
        source=inputs["source"],
        workbook=inputs["workbook"],
        rules=inputs["rules"],
        workbook_schema=inputs["workbook_schema"],
        scoping_answers=scoping_answers,
    )


def plant_invalid_credentials(monkeypatch):
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-invalid-cleared-by-adapter")
    monkeypatch.setenv("ANTHROPIC_AUTH_TOKEN", "invalid-cleared-by-adapter")
    monkeypatch.setenv("CODEX_API_KEY", "sk-invalid-cleared-by-adapter")
    monkeypatch.setenv("CODEX_ACCESS_TOKEN", "invalid-cleared-by-adapter")


def final_cell(workspace, cell):
    book = load_workbook(workspace.final_xlsx)
    try:
        return book[SHEET][cell].value
    finally:
        book.close()


def test_live_revision_and_bounded_rereview(inputs, monkeypatch):
    plant_invalid_credentials(monkeypatch)
    # A second source that genuinely contradicts the brief's start date.
    (inputs["source"] / NOTE_PATH).write_text(CONFLICTING_NOTE)

    fakes = FakeAgentRuntime({"filler": FILLER_FIXTURE, "reviewer": REVIEW_FIXTURE})
    claude = ClaudeCodeRuntime()
    codex = CodexRuntime()
    state = run_workflow(
        inputs=run_inputs(inputs, scoping_answers=inputs["scoping_answers"]),
        runs_root=inputs["runs_root"],
        runtimes={
            "filler": fakes,
            "reviewer": fakes,
            "revision": claude,
            "re_review": codex,
        },
    )
    workspace = Workspace(Path(state["workspace_path"]))

    # The confirmation-bias mitigation instructions are part of the
    # Revision invocation's prompt input (plan section 27).
    prompt_file, contract = CLAUDE_ROLES["revision"]
    assert contract is RevisionResult
    prompt = (PROMPTS_DIR / prompt_file).read_text()
    assert "Do not assume the Filler's original value is correct" in prompt
    assert "REBUT is reserved" in prompt

    # Revision received only the non-PASS findings plus the allowlist
    # (with the flagged rows' Notes cell authorized for note_append).
    revision_inputs = json.loads(
        (workspace.revision_outputs / "inputs.json").read_text()
    )
    assert {f["cell"] for f in revision_inputs["findings"]} == {"D2", "E2"}
    assert set(revision_inputs["proposals"]) == {"D2", "E2"}
    # Exactly the flagged cells plus their row's Notes cell — a PASS
    # cell leaking into the allowlist would fail here.
    allowed = {ref.split("!", 1)[1] for ref in revision_inputs["mutation_allowlist"]}
    assert allowed == {"D2", "E2", "F2"}

    revision = RevisionResult.model_validate_json(workspace.revision_json.read_text())
    decisions = {decision.cell: decision for decision in revision.decisions}
    assert set(decisions) == {"D2", "E2"}

    # The WARN without a recommendation leaves REBUT as the only legal
    # action, and the rebuttal flowed through exactly one live
    # re-review round over exactly that cell.
    assert decisions["E2"].action == "REBUT"
    verdicts = ReReviewResult.model_validate_json(workspace.re_review_json.read_text())
    assert [verdict.cell for verdict in verdicts.verdicts] == ["E2"]
    assert verdicts.verdicts[0].verdict in {"WITHDRAWN", "UPHELD"}

    # At least one decision exercised note_append; its text landed in
    # the row's Notes cell of the final workbook.
    noted = [d for d in revision.decisions if d.note_append]
    assert noted, "no revision decision exercised note_append"
    if any(d.cell == "D2" and d.action in {"CLEAR", "FIX"} for d in noted):
        notes_value = final_cell(workspace, "F2")
        assert notes_value and any(
            d.note_append in notes_value for d in noted if d.cell == "D2"
        )

    # PASS cells are frozen through revision and finalization.
    assert final_cell(workspace, "A2") == "PRJ-2008"
    assert workspace.final_xlsx.is_file()
    assert workspace.revision_log_md.is_file()


def test_full_live_run_on_the_sample_workspace(inputs, monkeypatch, capsys):
    plant_invalid_credentials(monkeypatch)
    # The CLI's own live wiring (ADR 0019) is what a real run uses.
    runtimes = build_runtimes("live")

    # Phase 1: live scoping pauses the run for answers.
    state = run_workflow(
        inputs=run_inputs(inputs),
        runs_root=inputs["runs_root"],
        runtimes=runtimes,
    )
    assert "__interrupt__" in state
    workspace = Workspace(Path(state["workspace_path"]))
    questions = json.loads(workspace.scoping_questions_json.read_text())
    assert questions["questions"]

    # Phase 2: answer and resume; every remaining role runs live.
    workspace.scoping_answers_md.write_text(
        "Answers to all scoping questions: one row per source folder"
        " (one project per folder), data rows starting at row 2. The"
        " provided folder set is the complete authoritative set. Apply"
        " the extraction rules exactly as written; no additional"
        " conventions.\n"
    )
    state = resume_workflow(
        run_id=state["run_id"],
        runs_root=inputs["runs_root"],
        runtimes=runtimes,
    )
    assert "__interrupt__" not in state
    workspace = Workspace(Path(state["workspace_path"]))

    # Every unconditional artifact of a completed run exists (plan
    # section 51).
    for artifact in (
        workspace.manifest_json,
        workspace.workbook_schema_json,
        workspace.review_policy_json,
        workspace.scoping_questions_json,
        workspace.scoping_questions_md,
        workspace.scoping_answers_md,
        workspace.extraction_json,
        workspace.validation_json,
        workspace.provenance_json,
        workspace.handoff_json,
        workspace.handoff_md,
        workspace.review_explorer_html,
        workspace.review_explorer_zh_html,
        workspace.review_json,
        workspace.review_md,
        workspace.draft_xlsx,
        workspace.final_xlsx,
        workspace.run_summary,
        workspace.audit_db,
    ):
        assert artifact.is_file(), f"missing artifact: {artifact.name}"

    # Path-dependent artifact families (plan section 51): whichever
    # routes the live verdicts took must have produced their artifacts.
    if state.get("revision_path") is not None:
        for artifact in (
            workspace.revision_json,
            workspace.revision_log_md,
            workspace.review_explorer_v2_html,
            workspace.review_explorer_zh_v2_html,
        ):
            assert artifact.is_file(), f"missing revision artifact: {artifact.name}"
    if state.get("re_review_path") is not None:
        assert workspace.re_review_json.is_file()
    if workspace.unresolved_json.is_file():
        assert workspace.human_review_json.is_file()
        assert workspace.human_review_md.is_file()

    # Provenance stays in sync with the final workbook: every entry's
    # recorded value matches the cell it describes.
    provenance = json.loads(workspace.provenance_json.read_text())
    assert provenance["entries"], "live run wrote no provenance entries"
    book = load_workbook(workspace.final_xlsx)
    try:
        sheet = book[SHEET]
        for entry in provenance["entries"]:
            cell_ref = entry["cell"].split("!", 1)[1]
            assert sheet[cell_ref].value == entry["value"], (
                f"provenance out of sync at {entry['cell']}"
            )
    finally:
        book.close()

    # PASS cells kept their originally proposed values. The Notes
    # column is exempt: it is the note_append companion channel, always
    # authorized for flagged rows regardless of its own verdict (plan
    # section 28).
    review = json.loads(workspace.review_json.read_text())
    extraction = json.loads(workspace.extraction_json.read_text())
    schema_config = json.loads(workspace.workbook_schema_json.read_text())
    sheet_config = next(s for s in schema_config["sheets"] if s["target"])
    notes_column = sheet_config["fields"][sheet_config["notes_field"]]["column"]
    proposed = {
        p["cell"]: p["value"]
        for p in extraction["proposals"]
        if p["status"] == "proposed"
    }
    pass_cells = [f["cell"] for f in review["findings"] if f["verdict"] == "PASS"]
    checked = 0
    for cell in pass_cells:
        if cell in proposed and cell.rstrip("0123456789") != notes_column:
            assert final_cell(workspace, cell) == proposed[cell], (
                f"PASS cell {cell} changed after review"
            )
            checked += 1
    assert checked, "no PASS cell could be checked against its proposal"

    err = capsys.readouterr().err
    assert "Claude Code auth: OAuth (subscription - best effort)" in err
    assert "Codex auth: ChatGPT subscription (auth.json)" in err
