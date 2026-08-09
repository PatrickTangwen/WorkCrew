"""Hostile-input suite for the deterministic workbook mutation layer
(ticket #4, plan sections 14, 28, 36, 37).

This is the agreed secondary test seam: the write boundary is unit
tested directly and exhaustively. Assertions inspect the reopened
workbook and the audit database — no mocking.
"""

import sqlite3

import pytest
from openpyxl import Workbook, load_workbook

from workflow_app.audit.db import AuditStore
from workflow_app.models import ReviewFinding, RevisionDecision
from workflow_app.workbook.mutations import (
    CellMutation,
    MutationConflictError,
    apply_mutations,
)
from workflow_app.workbook.safety import Allowlist
from workflow_app.workbook.schema import WorkbookSchema
from workflow_app.workflow.routing import compose_revision_mutations

SHEET = "7) Practicum Courses"

SCHEMA = WorkbookSchema.model_validate(
    {
        "sheets": [
            {
                "name": SHEET,
                "target": True,
                "notes_field": "Notes",
                "fields": {
                    "Project ID*": {
                        "type": "id",
                        "column": "A",
                        "pattern": r"^PRJ-\d{4}$",
                        "writable": True,
                    },
                    "Locked": {"type": "string", "column": "B"},
                    "Count": {"type": "number", "column": "C", "writable": True},
                    "Start Date": {"type": "date", "column": "D", "writable": True},
                    "Active": {"type": "boolean", "column": "E", "writable": True},
                    "Notes": {"type": "string", "column": "F", "writable": True},
                    "Main Issue Area(s)": {
                        "type": "controlled_vocabulary",
                        "column": "G",
                        "values": ["Healthcare", "Education"],
                        "writable": True,
                    },
                },
            }
        ]
    }
)

ALLOWED = Allowlist(
    [
        f"{SHEET}!A2",
        f"{SHEET}!B2",  # allowlisted but schema says not writable
        f"{SHEET}!C2",
        f"{SHEET}!D2",
        f"{SHEET}!E2",
        f"{SHEET}!F2",
        f"{SHEET}!G2",
        f"{SHEET}!H2",  # allowlisted but no field declared for column H
    ]
)

RUN_ID = "run-under-test"


@pytest.fixture
def draft(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET
    sheet["A2"] = "PRJ-0001"
    sheet["B2"] = "do not touch"
    sheet["C2"] = 1
    sheet["D2"] = "2026-01-01"
    sheet["E2"] = True
    sheet["G2"] = "Education"
    sheet["H2"] = "=SUM(1,2)"
    reference = workbook.create_sheet("Ref")
    reference["A1"] = "reference data"
    path = tmp_path / "draft.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def audit(tmp_path):
    store = AuditStore(tmp_path / "audit.sqlite")
    yield store
    store.close()


def mutation(cell, value, actor="filler", source="proposal-1", sheet=SHEET):
    return CellMutation(
        sheet=sheet, cell=cell, value=value, actor_role=actor, source_ref=source
    )


def apply_one(draft, audit, one):
    return apply_mutations(draft, [one], SCHEMA, ALLOWED, audit, RUN_ID)[0]


def cell_value(draft, cell, sheet=SHEET):
    return load_workbook(draft)[sheet][cell].value


def audit_rows(audit_path, status=None):
    query = (
        "SELECT sheet, cell, old_value, new_value, actor_role, source_ref,"
        " status, reason FROM mutations"
    )
    params = ()
    if status:
        query += " WHERE status = ?"
        params = (status,)
    with sqlite3.connect(audit_path) as conn:
        return conn.execute(query, params).fetchall()


# --- the happy path, fully audited --------------------------------------


def test_authorized_valid_mutation_is_applied_and_audited(draft, audit, tmp_path):
    outcome = apply_one(draft, audit, mutation("F2", "sourced note"))

    assert outcome.status == "applied"
    assert cell_value(draft, "F2") == "sourced note"

    rows = audit_rows(tmp_path / "audit.sqlite", "applied")
    assert rows == [
        (SHEET, "F2", "null", '"sourced note"', "filler", "proposal-1", "applied", None)
    ]


def test_overwriting_records_the_old_value(draft, audit, tmp_path):
    apply_one(draft, audit, mutation("G2", "Healthcare"))

    (row,) = audit_rows(tmp_path / "audit.sqlite", "applied")
    assert row[2] == '"Education"'  # old value preserved in the audit trail
    assert cell_value(draft, "G2") == "Healthcare"


# --- authorization ------------------------------------------------------


@pytest.mark.parametrize("actor", ["reviewer", "auditor", "", "FILLER"])
def test_roles_other_than_filler_and_revision_may_not_mutate(draft, audit, actor):
    outcome = apply_one(draft, audit, mutation("F2", "attempt", actor=actor))

    assert outcome.status == "rejected"
    assert "may not mutate" in outcome.reason
    assert cell_value(draft, "F2") is None


def test_cell_outside_allowlist_is_rejected_and_audited(draft, audit, tmp_path):
    outcome = apply_one(draft, audit, mutation("F9", "sneaky"))

    assert outcome.status == "rejected"
    assert "allowlist" in outcome.reason
    assert cell_value(draft, "F9") is None

    (row,) = audit_rows(tmp_path / "audit.sqlite", "rejected")
    assert row[1] == "F9" and "allowlist" in row[7]


def test_unwritable_column_is_rejected_even_if_allowlisted(draft, audit):
    outcome = apply_one(draft, audit, mutation("B2", "overwrite attempt"))

    assert outcome.status == "rejected"
    assert "writable" in outcome.reason
    assert cell_value(draft, "B2") == "do not touch"


def test_column_without_field_spec_is_rejected(draft, audit):
    outcome = apply_one(draft, audit, mutation("H2", "anything"))

    assert outcome.status == "rejected"
    assert "no field" in outcome.reason
    assert cell_value(draft, "H2") == "=SUM(1,2)"


def test_sheet_missing_from_workbook_is_rejected(draft, audit):
    outcome = apply_one(draft, audit, mutation("A1", "x", sheet="Ghost"))
    assert outcome.status == "rejected"
    assert "workbook" in outcome.reason


def test_sheet_not_described_by_schema_is_rejected(draft, audit):
    outcome = apply_one(draft, audit, mutation("A1", "x", sheet="Ref"))
    assert outcome.status == "rejected"
    assert "schema" in outcome.reason
    assert cell_value(draft, "A1", sheet="Ref") == "reference data"


@pytest.mark.parametrize("bad_cell", ["12G", "G", "2", "", "G12:H14", "G-2"])
def test_malformed_cell_addresses_are_rejected(draft, audit, bad_cell):
    outcome = apply_one(draft, audit, mutation(bad_cell, "x"))
    assert outcome.status == "rejected"
    assert "address" in outcome.reason


# --- each validator class blocks the write ------------------------------


@pytest.mark.parametrize(
    ("cell", "value"),
    [
        ("A2", "BAD-ID"),  # id pattern violation
        ("A2", 7),  # id type violation
        ("C2", "42"),  # number given a string
        ("C2", True),  # bool must not pass as number
        ("D2", "2026-99-99"),  # impossible date
        ("D2", 20260101),  # date given a number
        ("E2", "yes"),  # boolean given a string
        ("F2", 42),  # string given a number
        ("G2", "Sorcery"),  # outside controlled vocabulary
    ],
)
def test_validator_violations_reject_before_any_write(draft, audit, cell, value):
    before = cell_value(draft, cell)

    outcome = apply_one(draft, audit, mutation(cell, value))

    assert outcome.status == "rejected"
    assert cell_value(draft, cell) == before


# --- idempotency (plan section 37) --------------------------------------


def test_same_mutation_twice_audits_once_and_replays_the_value(draft, audit, tmp_path):
    first = apply_one(draft, audit, mutation("F2", "once"))
    second = apply_one(draft, audit, mutation("F2", "once"))

    assert first.status == "applied" and first.replayed is False
    # Downstream consumers see the same outcome as the first run; only
    # the audit trail knows it was a replay.
    assert second.status == "applied" and second.replayed is True
    assert cell_value(draft, "F2") == "once"
    assert len(audit_rows(tmp_path / "audit.sqlite", "applied")) == 1


def test_same_mutation_twice_within_one_batch_audits_once(draft, audit, tmp_path):
    outcomes = apply_mutations(
        draft,
        [mutation("F2", "once"), mutation("F2", "once")],
        SCHEMA,
        ALLOWED,
        audit,
        RUN_ID,
    )

    assert [outcome.status for outcome in outcomes] == ["applied", "applied"]
    assert [outcome.replayed for outcome in outcomes] == [False, True]
    assert cell_value(draft, "F2") == "once"
    assert len(audit_rows(tmp_path / "audit.sqlite", "applied")) == 1


def test_replay_after_workbook_reset_restores_the_value(draft, audit, tmp_path):
    template = draft.read_bytes()
    apply_one(draft, audit, mutation("F2", "once"))

    # A crash-resume re-copies the template: the audit remembers the
    # write but the workbook no longer holds it.
    draft.write_bytes(template)
    assert cell_value(draft, "F2") is None

    outcome = apply_one(draft, audit, mutation("F2", "once"))

    assert outcome.status == "applied" and outcome.replayed is True
    assert cell_value(draft, "F2") == "once"
    assert len(audit_rows(tmp_path / "audit.sqlite", "applied")) == 1


def test_intra_batch_same_source_different_value_conflicts_without_saving(
    draft, audit, tmp_path
):
    before = draft.read_bytes()
    with pytest.raises(MutationConflictError):
        apply_mutations(
            draft,
            [mutation("F2", "first"), mutation("F2", "second")],
            SCHEMA,
            ALLOWED,
            audit,
            RUN_ID,
        )
    # The conflict aborts the batch before any save or audit.
    assert draft.read_bytes() == before
    assert audit_rows(tmp_path / "audit.sqlite") == []


def test_replaying_same_source_with_different_value_is_a_conflict(draft, audit):
    apply_one(draft, audit, mutation("F2", "original"))
    with pytest.raises(MutationConflictError):
        apply_one(draft, audit, mutation("F2", "mutated story"))


def test_a_different_actor_may_rewrite_the_same_cell(draft, audit, tmp_path):
    fill = apply_one(draft, audit, mutation("F2", "first pass"))
    revise = apply_one(
        draft, audit, mutation("F2", "corrected", actor="revision", source="decision-9")
    )

    assert fill.status == revise.status == "applied"
    assert cell_value(draft, "F2") == "corrected"
    applied = audit_rows(tmp_path / "audit.sqlite", "applied")
    assert [(row[2], row[3], row[4]) for row in applied] == [
        ("null", '"first pass"', "filler"),
        ('"first pass"', '"corrected"', "revision"),
    ]


# --- batches ------------------------------------------------------------


def test_mixed_batch_applies_valid_and_rejects_invalid_independently(
    draft, audit, tmp_path
):
    outcomes = apply_mutations(
        draft,
        [mutation("F2", "kept"), mutation("F9", "sneaky"), mutation("G2", "Sorcery")],
        SCHEMA,
        ALLOWED,
        audit,
        RUN_ID,
    )

    assert [outcome.status for outcome in outcomes] == [
        "applied",
        "rejected",
        "rejected",
    ]
    assert cell_value(draft, "F2") == "kept"
    assert len(audit_rows(tmp_path / "audit.sqlite", "rejected")) == 2


def test_lowercase_cell_addresses_are_normalized(draft, audit):
    outcome = apply_one(draft, audit, mutation("f2", "normalized"))
    assert outcome.status == "applied"
    assert cell_value(draft, "F2") == "normalized"


# --- same-batch note composition through the real layers (ADR 0021) ------


def finding(cell, verdict, recommended=None):
    return ReviewFinding.model_validate(
        {
            "cell": cell,
            "verdict": verdict,
            "recommended_value": recommended,
            "evidence": [],
            "reviewer_comment": "because",
        }
    )


def decision(cell, action, proposed=None, note_append=None):
    return RevisionDecision.model_validate(
        {
            "cell": cell,
            "action": action,
            "proposed_value": proposed,
            "note_append": note_append,
            "evidence": [],
            "justification": "reasoned",
        }
    )


TWO_NOTE_FINDINGS = [finding("G2", "FAIL"), finding("D2", "FAIL")]
TWO_NOTE_DECISIONS = [
    decision("G2", "CLEAR", note_append="note one"),
    decision("D2", "FIX", proposed="2026-02-02", note_append="note two"),
]


def compose_and_apply(draft, audit, decisions=TWO_NOTE_DECISIONS):
    def find_prior(cell_ref, source_ref):
        return audit.find_applied_mutation(
            RUN_ID, SHEET, cell_ref, "revision", source_ref
        )

    mutations, _ = compose_revision_mutations(
        decisions,
        TWO_NOTE_FINDINGS,
        SCHEMA.target_sheet(),
        lambda cell_ref: cell_value(draft, cell_ref),
        find_prior,
    )
    return apply_mutations(draft, mutations, SCHEMA, ALLOWED, audit, RUN_ID)


def note_audit_rows(audit_path):
    applied = audit_rows(audit_path, "applied")
    return [(row[2], row[3]) for row in applied if row[1] == "F2"]


def test_two_note_batch_composes_through_the_real_layers(draft, audit, tmp_path):
    outcomes = compose_and_apply(draft, audit)

    assert [outcome.status for outcome in outcomes] == ["applied"] * 4
    assert cell_value(draft, "F2") == "note one\nnote two"
    assert note_audit_rows(tmp_path / "audit.sqlite") == [
        ("null", '"note one"'),
        ('"note one"', '"note one\\nnote two"'),
    ]


def test_two_note_replay_after_workbook_reset_is_identical(draft, audit, tmp_path):
    # ADR 0016's crash window: every mutation audited, the save lost.
    template = draft.read_bytes()
    compose_and_apply(draft, audit)
    draft.write_bytes(template)

    outcomes = compose_and_apply(draft, audit)

    assert [outcome.replayed for outcome in outcomes] == [True] * 4
    assert cell_value(draft, "F2") == "note one\nnote two"
    assert len(audit_rows(tmp_path / "audit.sqlite", "applied")) == 4


def test_partially_audited_two_note_batch_resumes_identically(draft, audit, tmp_path):
    # Crash in the middle of the audit loop: only the first decision's
    # mutations were audited and the save never happened. The resumed
    # batch composes the unaudited second note on the replayed first.
    template = draft.read_bytes()
    compose_and_apply(draft, audit, decisions=TWO_NOTE_DECISIONS[:1])
    draft.write_bytes(template)

    outcomes = compose_and_apply(draft, audit)

    assert [outcome.replayed for outcome in outcomes] == [True, True, False, False]
    assert cell_value(draft, "F2") == "note one\nnote two"
    assert note_audit_rows(tmp_path / "audit.sqlite") == [
        ("null", '"note one"'),
        ('"note one"', '"note one\\nnote two"'),
    ]


# --- the draft stays a sane workbook ------------------------------------


def test_untouched_cells_and_formulas_survive_save_and_reopen(draft, audit):
    apply_one(draft, audit, mutation("F2", "only this"))

    workbook = load_workbook(draft)
    sheet = workbook[SHEET]
    assert sheet["A2"].value == "PRJ-0001"
    assert sheet["B2"].value == "do not touch"
    assert sheet["E2"].value is True
    assert sheet["H2"].value == "=SUM(1,2)"  # formula, not its result
    assert workbook["Ref"]["A1"].value == "reference data"


def test_fully_rejected_batch_leaves_the_file_untouched(draft, audit):
    before = draft.read_bytes()
    apply_mutations(
        draft,
        [mutation("F9", "no"), mutation("B2", "no")],
        SCHEMA,
        ALLOWED,
        audit,
        RUN_ID,
    )
    assert draft.read_bytes() == before
