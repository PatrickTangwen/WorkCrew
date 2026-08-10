"""Kleister-Charity benchmark builder (ticket #13, plan section 42).

The builder is deterministic: fixed seed, sorted inputs, pure
derivation of the two WorkCrew-specific columns. Tests run against a
synthetic mini dev split shaped exactly like the real files.
"""

import json

import pytest

from workflow_app.benchmark.kleister import (
    build_benchmark,
    derive_fields,
    income_band,
    parse_split,
    sample_documents,
    unescape_text,
)
from workflow_app.workbook.schema import load_workbook_schema

KEYS = (
    "address__post_town address__postcode address__street_line charity_name"
    " charity_number income_annually_in_british_pounds report_date"
    " spending_annually_in_british_pounds"
)


def full_labels(number="1234567", income="500000.00"):
    return {
        "address__post_town": "BRISTOL",
        "address__postcode": "BS1_4DJ",
        "address__street_line": "12_HARBOUR_ROAD",
        "charity_name": "Harbour_Trust",
        "charity_number": number,
        "income_annually_in_british_pounds": income,
        "report_date": "2018-03-31",
        "spending_annually_in_british_pounds": "450000.00",
    }


def expected_line(labels):
    return " ".join(f"{key}={value}" for key, value in sorted(labels.items()))


def in_line(stem, text):
    escaped = text.replace("\n", "\\n").replace("\t", "\\t")
    return "\t".join([f"{stem}.pdf", KEYS, "djvu", "tesseract", "textract", escaped])


MINI_DOCS = [
    ("aaa1", "Report of Harbour Trust\nIncome: 500,000", full_labels()),
    ("bbb2", "Report two\nwith lines", full_labels(number="7654321")),
    ("ccc3", "Report three", full_labels(number="1111111")),
    # Missing income/spending/street -> natural expected-blank labels.
    (
        "ddd4",
        "Partial report",
        {
            "address__post_town": "LEEDS",
            "address__postcode": "LS1_1UR",
            "charity_name": "Leeds_Aid",
            "charity_number": "2222222",
            "report_date": "2019-12-31",
        },
    ),
    (
        "eee5",
        "Another partial",
        {
            "charity_name": "No_Address_Fund",
            "charity_number": "3333333",
            "report_date": "2020-01-31",
        },
    ),
    # Oversized document: must be excluded by the text cap.
    ("fff6", "x" * 5000, full_labels(number="4444444")),
]


@pytest.fixture
def dev_dir(tmp_path):
    split = tmp_path / "dev-0"
    split.mkdir()
    (split / "in.tsv").write_text(
        "\n".join(in_line(stem, text) for stem, text, _ in MINI_DOCS) + "\n"
    )
    (split / "expected.tsv").write_text(
        "\n".join(expected_line(labels) for _, _, labels in MINI_DOCS) + "\n"
    )
    return split


def build(dev_dir, output, **overrides):
    settings = {
        "sample_full": 2,
        "sample_partial": 1,
        "text_cap": 4000,
        "conflicts": 1,
        "seed": 7,
    }
    settings.update(overrides)
    return build_benchmark(dev_dir, output, **settings)


# --- text unescaping -----------------------------------------------------


def test_unescape_restores_newlines_and_tabs():
    assert unescape_text("a\\nb\\tc") == "a\nb\tc"


def test_unescape_preserves_other_backslash_sequences():
    # OCR noise like \D or \f is raw text, not an escape.
    assert unescape_text("a\\Db\\fc") == "a\\Db\\fc"


# --- split parsing -------------------------------------------------------


def test_parse_split_reads_sorted_docs_with_best_text(dev_dir):
    docs = parse_split(dev_dir)

    assert [doc["stem"] for doc in docs] == [
        "aaa1",
        "bbb2",
        "ccc3",
        "ddd4",
        "eee5",
        "fff6",
    ]
    assert docs[0]["text"] == "Report of Harbour Trust\nIncome: 500,000"
    assert docs[0]["labels"]["charity_number"] == "1234567"
    assert docs[3]["labels"].get("income_annually_in_british_pounds") is None


# --- derived columns -----------------------------------------------------


@pytest.mark.parametrize(
    ("income", "band"),
    [
        ("0.00", "Small"),
        ("249999.99", "Small"),
        ("250000.00", "Medium"),
        ("999999.99", "Medium"),
        ("1000000.00", "Large"),
        ("10348000.00", "Large"),
    ],
)
def test_income_band_thresholds(income, band):
    assert income_band(income) == band


def test_derive_fields_builds_the_two_derived_columns():
    fields = derive_fields(full_labels())

    assert fields["Charity ID*"] == "CHA-1234567"
    assert fields["Registration Number"] == "1234567"
    assert fields["Income Size Band"] == "Medium"
    assert fields["Annual Income GBP"] == "500000.00"
    assert fields["Post Town"] == "BRISTOL"


def test_derive_fields_blanks_derived_columns_without_their_source():
    fields = derive_fields({"charity_name": "X", "report_date": "2020-01-01"})

    assert fields["Charity ID*"] is None
    assert fields["Income Size Band"] is None
    assert fields["Annual Income GBP"] is None
    assert fields["Charity Name"] == "X"


# --- sampling ------------------------------------------------------------


def test_sampling_is_deterministic_and_respects_cap_and_strata(dev_dir):
    docs = parse_split(dev_dir)

    once = sample_documents(
        docs, sample_full=2, sample_partial=1, text_cap=4000, seed=7
    )
    again = sample_documents(
        docs, sample_full=2, sample_partial=1, text_cap=4000, seed=7
    )

    assert [doc["stem"] for doc in once] == [doc["stem"] for doc in again]
    assert len(once) == 3
    stems = [doc["stem"] for doc in once]
    assert stems == sorted(stems)
    assert "fff6" not in stems  # over the text cap
    partial = [doc for doc in once if len(doc["labels"]) < 8]
    assert len(partial) == 1


def test_sampling_refuses_an_undersized_population(dev_dir):
    docs = parse_split(dev_dir)
    with pytest.raises(ValueError, match="sample"):
        sample_documents(docs, sample_full=10, sample_partial=1, text_cap=4000, seed=7)


# --- full build ----------------------------------------------------------


def test_build_writes_folders_labels_schema_and_rules(dev_dir, tmp_path):
    output = tmp_path / "bench"
    summary = build(dev_dir, output)

    labels = json.loads((output / "labels.json").read_text())
    assert labels["sheet"] == "Charity Reports"
    assert len(labels["rows"]) == 3
    assert summary["documents"] == 3

    for index, row in enumerate(labels["rows"]):
        assert row["row"] == 2 + index
        folder = output / "source" / row["folder"]
        assert (folder / "report.txt").is_file()
        assert row["document"] == f"{row['folder']}/report.txt"

    # Rows follow sorted folder order; scoping answers pin the mapping.
    folders = [row["folder"] for row in labels["rows"]]
    assert folders == sorted(folders)
    answers = (output / "scoping_answers.md").read_text()
    assert "alphabetical" in answers

    # ADR 0015: benchmark schema declares the explorer annotations.
    schema = load_workbook_schema(output / "workbook_schema.json")
    sheet = schema.target_sheet()
    assert sheet.title_field is not None
    assert sheet.overview_fields
    assert all(field.gloss_zh for field in sheet.fields.values())
    assert sheet.notes_field == "Notes"

    # All three rule documents land in the single rules file (ADR 0032).
    rules_text = (output / "rules.md").read_text()
    assert "CHA-" in rules_text
    assert "250,000" in rules_text

    # License note is recorded with the benchmark (issue #13 comment).
    readme = (output / "README.md").read_text()
    assert "license" in readme.lower()

    from openpyxl import load_workbook

    headers = next(
        load_workbook(output / "template.xlsx")["Charity Reports"].iter_rows(
            max_row=1, values_only=True
        )
    )
    assert headers[0] == "Charity ID*"
    assert "Notes" in headers


def test_build_marks_expected_blanks_for_missing_fields(dev_dir, tmp_path):
    output = tmp_path / "bench"
    build(dev_dir, output, conflicts=0)

    labels = json.loads((output / "labels.json").read_text())
    partial_rows = [
        row
        for row in labels["rows"]
        if row["fields"]["Annual Income GBP"]["status"] == "blank"
    ]
    assert partial_rows
    row = partial_rows[0]
    assert row["fields"]["Annual Income GBP"]["expected_value"] is None
    assert row["fields"]["Income Size Band"]["status"] == "blank"
    assert row["fields"]["Charity Name"]["status"] == "expected"


def test_conflict_injection_writes_second_source_and_unresolved_labels(
    dev_dir, tmp_path
):
    output = tmp_path / "bench"
    build(dev_dir, output, conflicts=1)

    labels = json.loads((output / "labels.json").read_text())
    conflicted = [row for row in labels["rows"] if row["conflict"] is not None]
    assert len(conflicted) == 1
    row = conflicted[0]

    extract = output / "source" / row["folder"] / row["conflict"]["file"]
    assert extract.is_file()
    assert row["conflict"]["value"] in extract.read_text()

    income = row["fields"]["Annual Income GBP"]
    band = row["fields"]["Income Size Band"]
    assert income["status"] == "unresolved" and income["expected_value"] is None
    assert band["status"] == "unresolved" and band["expected_value"] is None
    # The conflicting copy genuinely disagrees with the report label.
    assert (
        row["conflict"]["value"] != full_labels()["income_annually_in_british_pounds"]
    )


def test_build_is_reproducible_byte_for_byte(dev_dir, tmp_path):
    build(dev_dir, tmp_path / "one")
    build(dev_dir, tmp_path / "two")

    # Every artifact, including the xlsx (whose zip/docProps timestamps
    # are pinned), rebuilds byte-identically.
    for name in ("labels.json", "workbook_schema.json", "template.xlsx"):
        assert (tmp_path / "one" / name).read_bytes() == (
            tmp_path / "two" / name
        ).read_bytes(), name
