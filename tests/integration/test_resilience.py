"""Crash-resume and retry resilience (ticket #9, plan section 37).

Seam: the engine entries with scripted runtimes injected. A "kill" step
raises KeyboardInterrupt — the injectable equivalent of the process
dying mid-stage: it must never be retried and leaves the checkpoint at
the last completed node. Assertions compare artifacts, workbook cells,
and the audit database against an uninterrupted baseline run.
"""

import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook

from workflow_app.runtimes.base import AgentResult
from workflow_app.workflow.engine import resume_workflow, run_workflow
from workflow_app.workspace import RunInputs

SHEET = "7) Practicum Courses"
BRIEF = "India 2008/Project_Brief.txt"


class ScriptedRuntime:
    """Replays a per-role sequence of behaviors: "kill" raises
    KeyboardInterrupt, "error" returns a failed AgentResult, "raise"
    throws a runtime exception, and a dict is a successful output. The
    last step repeats for any further calls."""

    name = "fake"

    def __init__(self, scripts):
        self._scripts = {role: list(steps) for role, steps in scripts.items()}
        self.calls = {role: 0 for role in scripts}

    def run(self, request):
        self.calls[request.role] += 1
        steps = self._scripts[request.role]
        step = steps.pop(0) if len(steps) > 1 else steps[0]
        if step == "kill":
            raise KeyboardInterrupt("injected kill")
        if step == "error":
            return AgentResult(status="error", error="injected process failure")
        if step == "raise":
            raise RuntimeError("injected invocation failure")
        return AgentResult(status="ok", output=step)


def evidence(text):
    return {
        "source_file": BRIEF,
        "source_location": "page 1",
        "evidence_text": text,
        "evidence_type": "direct",
    }


SCOPING_OUTPUT = {"questions": [{"id": "Q1", "question": "One row per folder?"}]}

FILLER_OUTPUT = {
    "proposals": [
        {
            "sheet": SHEET,
            "row": 2,
            "column_name": "Notes",
            "cell": "F2",
            "value": "First note.",
            "evidence": [evidence("Stated in the brief.")],
            "rules_applied": [],
            "confidence": "high",
            "status": "proposed",
        },
        {
            "sheet": SHEET,
            "row": 2,
            "column_name": "Project ID*",
            "cell": "A2",
            "value": "PRJ-0001",
            "evidence": [evidence("Constructed from the folder name.")],
            "rules_applied": [],
            "confidence": "medium",
            "status": "proposed",
        },
    ]
}

REVIEW_OUTPUT = {
    "findings": [
        {
            "cell": "F2",
            "verdict": "WARN",
            "recommended_value": "Better community note.",
            "evidence": [evidence("Report names the community.")],
            "reviewer_comment": "Note misses the community.",
        },
        {
            "cell": "A2",
            "verdict": "WARN",
            "recommended_value": "PRJ-0002",
            "evidence": [evidence("Register suggests a second project.")],
            "reviewer_comment": "ID may collide.",
        },
    ]
}

FILLER_OUTPUT["proposals"].append(
    {
        "sheet": SHEET,
        "row": 4,
        "column_name": "Main Issue Area(s)",
        "cell": "G4",
        "value": "Education",
        "evidence": [evidence("Mentioned once in the archive notes.")],
        "rules_applied": [],
        "confidence": "medium",
        "status": "proposed",
    }
)
REVIEW_OUTPUT["findings"].append(
    {
        "cell": "G4",
        "verdict": "FAIL",
        "recommended_value": None,
        "evidence": [evidence("No source supports an issue area here.")],
        "reviewer_comment": "Unsupported issue area.",
    }
)

REVISION_OUTPUT = {
    "decisions": [
        {
            "cell": "F2",
            "action": "ACCEPT",
            "proposed_value": None,
            "note_append": None,
            "evidence": [evidence("Reviewer evidence is stronger.")],
            "justification": "Adopting the recommendation.",
        },
        {
            "cell": "A2",
            "action": "REBUT",
            "proposed_value": None,
            "note_append": None,
            "evidence": [evidence("Brief explicitly names PRJ-0001.")],
            "justification": "The register entry is a different program.",
        },
        {
            "cell": "G4",
            "action": "CLEAR",
            "proposed_value": None,
            "note_append": "Issue area removed; unsupported by sources.",
            "evidence": [evidence("Re-checked; nothing supports it.")],
            "justification": "Cannot be determined from sources.",
        },
    ]
}

RE_REVIEW_OUTPUT = {
    "verdicts": [
        {"cell": "A2", "verdict": "UPHELD", "reviewer_comment": "Register stands."}
    ]
}

OK_SCRIPTS = {
    "scoping": [SCOPING_OUTPUT],
    "filler": [FILLER_OUTPUT],
    "reviewer": [REVIEW_OUTPUT],
    "revision": [REVISION_OUTPUT],
    "re_review": [RE_REVIEW_OUTPUT],
}


def make_runtimes(overrides=None):
    scripts = {role: list(steps) for role, steps in OK_SCRIPTS.items()}
    for role, steps in (overrides or {}).items():
        scripts[role] = list(steps)
    runtime = ScriptedRuntime(scripts)
    return {role: runtime for role in scripts}, runtime


def run_inputs(inputs, **overrides):
    values = {
        "source": inputs["source"],
        "workbook": inputs["workbook"],
        "rules": inputs["rules"],
        "workbook_schema": inputs["workbook_schema"],
        "scoping_answers": inputs["scoping_answers"],
        **overrides,
    }
    return RunInputs(**values)


def start_run(inputs, runtimes, runs_root=None):
    return run_workflow(
        inputs=run_inputs(inputs),
        runs_root=runs_root or inputs["runs_root"],
        runtimes=runtimes,
    )


def only_run_id(runs_root):
    (workspace,) = list(Path(runs_root).iterdir())
    return workspace.name


def sheet_cells(workbook_path):
    sheet = load_workbook(workbook_path)[SHEET]
    return {
        f"{cell.coordinate}": cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }


def normalized_artifact(workspace, name, run_id):
    text = (Path(workspace) / "artifacts" / name).read_text()
    return json.loads(text.replace(run_id, "RUN"))


def audit_query(workspace, query, params=()):
    with sqlite3.connect(Path(workspace) / "state/audit.sqlite") as conn:
        return conn.execute(query, params).fetchall()


COMPARED_ARTIFACTS = (
    "extraction.json",
    "review.json",
    "revision.json",
    "re_review.json",
    "provenance.json",
    "human_review.json",
    "unresolved.json",
)


@pytest.mark.parametrize("kill_role", ["filler", "reviewer", "revision", "re_review"])
def test_kill_and_resume_matches_an_uninterrupted_run(inputs, tmp_path, kill_role):
    baseline_root = tmp_path / "baseline-runs"
    baseline_runtimes, _ = make_runtimes()
    baseline_state = start_run(inputs, baseline_runtimes, runs_root=baseline_root)
    baseline = baseline_root / baseline_state["run_id"]

    ok_output = OK_SCRIPTS[kill_role][0]
    runtimes, _ = make_runtimes({kill_role: ["kill", ok_output]})
    with pytest.raises(KeyboardInterrupt):
        start_run(inputs, runtimes)
    run_id = only_run_id(inputs["runs_root"])

    state = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    workspace = inputs["runs_root"] / run_id
    assert "__interrupt__" not in state

    # Same final workbook cells and the same artifact contents.
    assert sheet_cells(workspace / "output/final.xlsx") == sheet_cells(
        baseline / "output/final.xlsx"
    )
    for name in COMPARED_ARTIFACTS:
        assert normalized_artifact(workspace, name, run_id) == normalized_artifact(
            baseline, name, baseline_state["run_id"]
        ), name

    # No duplicate applied mutations in the audit trail.
    duplicates = audit_query(
        workspace,
        "SELECT sheet, cell, actor_role, source_ref, COUNT(*) FROM mutations"
        " WHERE status = 'applied' GROUP BY sheet, cell, actor_role, source_ref"
        " HAVING COUNT(*) > 1",
    )
    assert duplicates == []
    assert len(
        audit_query(workspace, "SELECT * FROM mutations WHERE status='applied'")
    ) == len(audit_query(baseline, "SELECT * FROM mutations WHERE status='applied'"))


def stage_row(workspace, stage):
    rows = audit_query(
        workspace,
        "SELECT status, retry_count, failure FROM stages"
        " WHERE stage = ? ORDER BY id DESC LIMIT 1",
        (stage,),
    )
    assert rows, f"no audit row for stage {stage}"
    return rows[0]


def test_transient_failures_retry_twice_then_succeed(inputs):
    runtimes, runtime = make_runtimes({"reviewer": ["error", "raise", REVIEW_OUTPUT]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    assert (workspace / "output/final.xlsx").is_file()
    assert runtime.calls["reviewer"] == 3
    status, retry_count, failure = stage_row(workspace, "CODEX_REVIEW")
    assert (status, retry_count, failure) == ("completed", 2, None)


def test_contract_violating_output_is_retried(inputs):
    malformed = {"findings": [{"cell": "F2"}]}
    runtimes, runtime = make_runtimes({"reviewer": [malformed, REVIEW_OUTPUT]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    assert (workspace / "output/final.xlsx").is_file()
    assert runtime.calls["reviewer"] == 2
    status, retry_count, failure = stage_row(workspace, "CODEX_REVIEW")
    assert (status, retry_count, failure) == ("completed", 1, None)


def run_status(workspace):
    (row,) = audit_query(workspace, "SELECT status FROM runs")
    return row[0]


def human_review_items(workspace):
    payload = json.loads((workspace / "artifacts/human_review.json").read_text())
    return {item["cell"]: item for item in payload["items"]}


def test_reviewer_triple_failure_escalates_all_written_cells(inputs):
    runtimes, runtime = make_runtimes({"reviewer": ["error", "raise", "error"]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    assert runtime.calls["reviewer"] == 3
    # Nothing was reviewed, so nothing may pretend to be all-clear:
    # every agent-written cell escalates to human review.
    assert (workspace / "output/final.xlsx").is_file()
    assert run_status(workspace) == "completed"
    assert runtime.calls["revision"] == 0

    items = human_review_items(workspace)
    assert set(items) == {"F2", "A2", "G4"}
    for item in items.values():
        assert "review" in item["reason"]
        assert item["reviewer"] is None

    status, retry_count, failure = stage_row(workspace, "CODEX_REVIEW")
    assert (status, retry_count, failure) == (
        "completed",
        2,
        "runtime_process_failure",
    )


def test_revision_triple_failure_marks_all_findings_unresolved(inputs):
    runtimes, runtime = make_runtimes({"revision": ["raise", "error", "raise"]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    assert runtime.calls["revision"] == 3
    assert runtime.calls["re_review"] == 0
    assert run_status(workspace) == "completed"

    items = human_review_items(workspace)
    assert set(items) == {"F2", "A2", "G4"}
    for item in items.values():
        assert item["revision"] is None
        assert "no revision decision" in item["reason"]

    status, retry_count, failure = stage_row(workspace, "CLAUDE_REVISE")
    assert (status, retry_count, failure) == ("completed", 2, "invocation_failure")


def test_re_review_triple_failure_marks_rebuttals_unresolved(inputs):
    runtimes, runtime = make_runtimes({"re_review": ["error", "error", "error"]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    assert runtime.calls["re_review"] == 3
    assert run_status(workspace) == "completed"

    items = human_review_items(workspace)
    # Only the rebutted cell is unadjudicated; the ACCEPT went through.
    assert set(items) == {"A2"}
    assert "re-review" in items["A2"]["reason"]

    status, retry_count, failure = stage_row(workspace, "CODEX_REREVIEW")
    assert (status, retry_count, failure) == (
        "completed",
        2,
        "runtime_process_failure",
    )


def test_filler_triple_failure_aborts_and_a_later_resume_retries(inputs):
    runtimes, runtime = make_runtimes(
        {"filler": ["error", "raise", "error", FILLER_OUTPUT]}
    )
    with pytest.raises(Exception, match="filler failed after 3 attempts"):
        start_run(inputs, runtimes)
    run_id = only_run_id(inputs["runs_root"])
    workspace = inputs["runs_root"] / run_id

    assert runtime.calls["filler"] == 3
    assert run_status(workspace) == "failed"
    status, retry_count, failure = stage_row(workspace, "CLAUDE_FILL")
    assert (status, retry_count, failure) == ("failed", 2, "runtime_process_failure")

    # The checkpoint survives: a later resume gets fresh attempts.
    state = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    assert "__interrupt__" not in state
    assert (workspace / "output/final.xlsx").is_file()
    assert run_status(workspace) == "completed"
    assert runtime.calls["filler"] == 4


def test_deterministic_failures_are_not_retried(inputs):
    bad_review = {
        "findings": [
            {
                "cell": "NOT-A-CELL",
                "verdict": "WARN",
                "recommended_value": None,
                "evidence": [evidence("Broken addressing.")],
                "reviewer_comment": "Malformed.",
            }
        ]
    }
    runtimes, runtime = make_runtimes({"reviewer": [bad_review]})
    with pytest.raises(ValueError, match="malformed cell"):
        start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / only_run_id(inputs["runs_root"])

    assert runtime.calls["reviewer"] == 1
    assert run_status(workspace) == "failed"
    status, retry_count, failure = stage_row(workspace, "CODEX_REVIEW")
    assert (status, retry_count, failure) == ("failed", 0, "deterministic")


def test_kill_inside_the_apply_node_before_save_resumes_identically(
    inputs, tmp_path, monkeypatch
):
    # The crash window the audit-before-save ordering exists for: the
    # revision mutations are audited, the workbook save never happens,
    # and the checkpoint still points at APPLY_ALLOWED_REVISIONS. The
    # resumed run must replay to the same result — in particular the
    # note_append must not double-apply.
    from workflow_app.workbook import writer

    baseline_root = tmp_path / "baseline-runs"
    baseline_runtimes, _ = make_runtimes()
    baseline_state = start_run(inputs, baseline_runtimes, runs_root=baseline_root)
    baseline = baseline_root / baseline_state["run_id"]

    real_save = writer.save_draft
    armed = {"kills_left": 1}

    def killing_save(workbook, path):
        # The fill-stage save passes; the revision-stage save is killed.
        if "draft" in Path(path).name and armed["kills_left"] and armed["revision"]:
            armed["kills_left"] -= 1
            raise KeyboardInterrupt("injected kill before save")
        real_save(workbook, path)

    armed["revision"] = False

    runtimes, _ = make_runtimes()

    # Arm the kill only once the revision decisions exist on disk —
    # i.e. the save happening inside APPLY_ALLOWED_REVISIONS.
    class ArmingRuntime:
        name = "fake"

        def __init__(self, inner):
            self._inner = inner

        def run(self, request):
            result = self._inner.run(request)
            if request.role == "revision":
                armed["revision"] = True
            return result

    inner = runtimes["revision"]
    arming = ArmingRuntime(inner)
    wrapped = {role: arming for role in runtimes}
    monkeypatch.setattr(writer, "save_draft", killing_save)

    with pytest.raises(KeyboardInterrupt):
        run_workflow(
            inputs=run_inputs(inputs),
            runs_root=inputs["runs_root"],
            runtimes=wrapped,
        )
    run_id = only_run_id(inputs["runs_root"])
    workspace = inputs["runs_root"] / run_id

    state = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=wrapped
    )
    assert "__interrupt__" not in state

    assert sheet_cells(workspace / "output/final.xlsx") == sheet_cells(
        baseline / "output/final.xlsx"
    )
    for name in COMPARED_ARTIFACTS:
        assert normalized_artifact(workspace, name, run_id) == normalized_artifact(
            baseline, name, baseline_state["run_id"]
        ), name
    duplicates = audit_query(
        workspace,
        "SELECT sheet, cell, actor_role, source_ref, COUNT(*) FROM mutations"
        " WHERE status = 'applied' GROUP BY sheet, cell, actor_role, source_ref"
        " HAVING COUNT(*) > 1",
    )
    assert duplicates == []


def test_scoping_triple_failure_aborts_then_resume_retries_and_pauses(inputs):
    runtimes, runtime = make_runtimes(
        {"scoping": ["error", "raise", "error", SCOPING_OUTPUT]}
    )
    with pytest.raises(Exception, match="scoping failed after 3 attempts"):
        run_workflow(
            inputs=run_inputs(inputs, scoping_answers=None),
            runs_root=inputs["runs_root"],
            runtimes=runtimes,
        )
    run_id = only_run_id(inputs["runs_root"])
    workspace = inputs["runs_root"] / run_id
    assert run_status(workspace) == "failed"
    status, retry_count, failure = stage_row(workspace, "CLAUDE_SCOPE")
    assert (status, retry_count, failure) == ("failed", 2, "runtime_process_failure")

    paused = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    assert "__interrupt__" in paused
    assert run_status(workspace) == "paused"
    assert runtime.calls["scoping"] == 4

    (workspace / "artifacts/scoping_answers.md").write_text("Q1: yes.")
    state = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    assert "__interrupt__" not in state
    assert (workspace / "output/final.xlsx").is_file()


def test_resuming_a_completed_run_is_refused(inputs):
    runtimes, _ = make_runtimes()
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    with pytest.raises(ValueError, match="already completed"):
        resume_workflow(
            run_id=state["run_id"],
            runs_root=inputs["runs_root"],
            runtimes=runtimes,
        )
    assert run_status(workspace) == "completed"


def test_retries_leave_classified_events_even_after_success(inputs):
    runtimes, _ = make_runtimes({"reviewer": ["error", "raise", REVIEW_OUTPUT]})
    state = start_run(inputs, runtimes)
    workspace = inputs["runs_root"] / state["run_id"]

    events = audit_query(
        workspace,
        "SELECT payload FROM events WHERE kind = 'stage_retry' ORDER BY id",
    )
    payloads = [json.loads(payload) for (payload,) in events]
    assert [p["classification"] for p in payloads] == [
        "runtime_process_failure",
        "invocation_failure",
    ]
    assert all(p["stage"] == "CODEX_REVIEW" for p in payloads)


def test_kill_during_scoping_then_resume_pauses_and_completes(inputs):
    runtimes, _ = make_runtimes({"scoping": ["kill", SCOPING_OUTPUT]})
    with pytest.raises(KeyboardInterrupt):
        run_workflow(
            inputs=run_inputs(inputs, scoping_answers=None),
            runs_root=inputs["runs_root"],
            runtimes=runtimes,
        )
    run_id = only_run_id(inputs["runs_root"])
    workspace = inputs["runs_root"] / run_id

    # First resume re-runs the scoping pass and pauses for answers.
    paused = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    assert "__interrupt__" in paused
    (workspace / "artifacts/scoping_answers.md").write_text("Q1: yes.")

    state = resume_workflow(
        run_id=run_id, runs_root=inputs["runs_root"], runtimes=runtimes
    )
    assert "__interrupt__" not in state
    assert (workspace / "output/final.xlsx").is_file()
    # The scoping answers were ingested exactly once.
    events = audit_query(
        workspace,
        "SELECT COUNT(*) FROM events WHERE kind = 'scoping_answers_received'",
    )
    assert events == [(1,)]
