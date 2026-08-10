"""Cross-domain contracts for the schema- and rule-driven workflow."""

import json
import re
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from workflow_app.runtimes.fake import FakeAgentRuntime
from workflow_app.workflow.engine import run_workflow
from workflow_app.workspace import RunInputs

CASES = [
    {
        "id": "invoice_to_ap",
        "folder": "vendor_batch_07",
        "sheet": "Accounts Payable Intake",
        "identity": ("Invoice Code", "B", "INV-007"),
        "source_identity": "007",
        "identity_rule": "Canonical Invoice Code: prefix the source invoice number with INV-.",
        "conflict": ("Gross Total", "E"),
        "source_text": "Invoice number 007 lists totals of 820 and 870 in two summaries.",
    },
    {
        "id": "application_to_crm",
        "folder": "candidate_19",
        "sheet": "Applicant Register",
        "identity": ("Applicant Key", "C", "APP-019"),
        "source_identity": "019",
        "identity_rule": "Canonical Applicant Key: prefix the source applicant number with APP-.",
        "conflict": ("Requested Support", "F"),
        "source_text": "Applicant number 019 requests 40 hours; the addendum requests 60.",
    },
]


def evidence(source_file, text, evidence_type="direct", location="line 1"):
    return {
        "source_file": source_file,
        "source_location": location,
        "evidence_text": text,
        "evidence_type": evidence_type,
    }


def stage_names(workspace, run_id):
    with sqlite3.connect(workspace / "state/audit.sqlite") as conn:
        rows = conn.execute(
            "SELECT stage FROM stages WHERE run_id = ? ORDER BY id", (run_id,)
        ).fetchall()
    return [row[0] for row in rows]


def explorer_data(path):
    match = re.search(r"const DATA = (.*);$", path.read_text(), re.MULTILINE)
    assert match is not None
    return json.loads(match.group(1))


@pytest.mark.parametrize("case", CASES, ids=[item["id"] for item in CASES])
def test_workflow_invariants_survive_domain_and_layout_changes(tmp_path, case):
    source = tmp_path / "source"
    folder = source / case["folder"]
    folder.mkdir(parents=True)
    source_file = f"{case['folder']}/record.txt"
    (folder / "record.txt").write_text(case["source_text"])

    identity_name, identity_column, identity_value = case["identity"]
    conflict_name, conflict_column = case["conflict"]
    workbook = tmp_path / "template.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = case["sheet"]
    sheet[f"{identity_column}1"] = identity_name
    sheet[f"{conflict_column}1"] = conflict_name
    book.save(workbook)

    schema = tmp_path / "workbook_schema.json"
    schema.write_text(
        json.dumps(
            {
                "sheets": [
                    {
                        "name": case["sheet"],
                        "target": True,
                        "fields": {
                            identity_name: {
                                "column": identity_column,
                                "value_kind": "constructed",
                                "writable": True,
                            },
                            conflict_name: {
                                "column": conflict_column,
                                "type": "number",
                                "writable": True,
                            },
                        },
                    }
                ]
            }
        )
    )
    rules = tmp_path / "rules"
    rules.mkdir()
    (rules / "field_rules.md").write_text(
        "# Canonical identifier\n\n"
        + case["identity_rule"]
        + "\n\nPreserve unresolved conflicts.\n"
    )
    answers = tmp_path / "scoping_answers.md"
    answers.write_text(f"Row 2 maps to {case['folder']}.\n")
    policy = tmp_path / "review_policy.yaml"
    policy.write_text("review:\n  coverage: full\n")

    identity_cell = f"{identity_column}2"
    conflict_cell = f"{conflict_column}2"
    fill = {
        "proposals": [
            {
                "sheet": case["sheet"],
                "row": 2,
                "column_name": identity_name,
                "cell": identity_cell,
                "value": identity_value,
                "evidence": [
                    evidence(
                        source_file,
                        f"The record states base identifier {case['source_identity']}.",
                    ),
                    evidence(
                        "rules/field_rules.md",
                        case["identity_rule"],
                        evidence_type="rule",
                        location="Canonical identifier",
                    ),
                ],
                "rules_applied": ["field_rules.md#canonical-identifier"],
                "confidence": "medium",
                "status": "proposed",
                "notes": "Constructed from the source identifier under the local canonical rule.",
            },
            {
                "sheet": case["sheet"],
                "row": 2,
                "column_name": conflict_name,
                "cell": conflict_cell,
                "value": None,
                "evidence": [
                    evidence(source_file, "Two authoritative passages disagree.")
                ],
                "rules_applied": [],
                "confidence": None,
                "status": "conflict",
                "notes": "The competing values cannot be reconciled.",
            },
        ]
    }
    review = {
        "findings": [
            {
                "cell": cell,
                "verdict": "PASS",
                "evidence": [evidence(source_file, "Checked against the record.")],
                "reviewer_comment": "The current workbook representation is correct.",
            }
            for cell in (identity_cell, conflict_cell)
        ]
    }
    fake = FakeAgentRuntime({"filler": fill, "reviewer": review})

    state = run_workflow(
        inputs=RunInputs(
            source=source,
            workbook=workbook,
            rules=rules,
            workbook_schema=schema,
            scoping_answers=answers,
            review_policy=policy,
        ),
        runs_root=tmp_path / "runs",
        runtimes={"filler": fake, "reviewer": fake},
    )

    workspace = Path(state["workspace_path"])
    final = load_workbook(workspace / "output/final.xlsx")[case["sheet"]]
    assert final[identity_cell].value == identity_value
    assert final[conflict_cell].value is None
    assert "CLAUDE_REVISE" not in stage_names(workspace, state["run_id"])

    reviewer_inputs = json.loads(
        (workspace / "agent_outputs/reviewer/inputs.json").read_text()
    )
    assert reviewer_inputs["review_targets"] == [
        {"cell": identity_cell, "reason": "full coverage"},
        {"cell": conflict_cell, "reason": "full coverage"},
    ]
    assert json.loads((workspace / "artifacts/unresolved.json").read_text()) == {
        "cells": [
            {
                "cell": conflict_cell,
                "reason": "protected source conflict requires human review",
            }
        ]
    }
    handoff = json.loads((workspace / "artifacts/handoff.json").read_text())
    assert [item["cell"] for item in handoff["decision_records"]] == [
        f"{case['sheet']}!{identity_cell}",
        f"{case['sheet']}!{conflict_cell}",
    ]
    assert handoff["decision_records"][0]["evidence"][0]["source_file"] == source_file
    identity_record = handoff["decision_records"][0]
    assert [item["evidence_type"] for item in identity_record["evidence"]] == [
        "direct",
        "rule",
    ]
    assert identity_record["rules_applied"] == ["field_rules.md#canonical-identifier"]
    assert identity_record["review_note"] == (
        "Constructed from the source identifier under the local canonical rule."
    )

    v1 = explorer_data(workspace / "artifacts/review_explorer.html")
    v2 = explorer_data(workspace / "artifacts/review_explorer_v2.html")
    (v1_row,) = v1["rows"]
    (v2_row,) = v2["rows"]
    v1_fields = {item["name"]: item for item in v1_row["fields"]}
    v2_fields = {item["name"]: item for item in v2_row["fields"]}

    assert v1_fields[identity_name]["proposal"]["rules_applied"] == [
        "field_rules.md#canonical-identifier"
    ]
    assert [
        item["type"] for item in v1_fields[identity_name]["proposal"]["evidence"]
    ] == ["direct", "rule"]
    assert v1_fields[conflict_name]["proposal"]["status"] == "conflict"
    assert v2_fields[conflict_name]["review"]["verdict"] == "PASS"
    assert v2_fields[conflict_name]["unresolved_reason"] == (
        "protected source conflict requires human review"
    )
    assert v2["review_cycle"]["unresolved_count"] == 1
