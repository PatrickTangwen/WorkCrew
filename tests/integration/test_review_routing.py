"""Review -> revision -> re-review -> human-fallback routing (ticket #6).

Primary seam: engine entry with fakes for every role. The fixture
distribution exercises every route: PASS (frozen), WARN->ACCEPT,
WARN->REBUT->WITHDRAWN, FAIL->CLEAR with note_append, and
WARN->REBUT->UPHELD escalating to human review. Assertions inspect the
final workbook, artifacts, and audit only.
"""

import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.integration.conftest import scoping_fixture
from workflow_app.runtimes.fake import FakeAgentRuntime
from workflow_app.workflow.engine import run_workflow
from workflow_app.workspace import RunInputs

SHEET = "7) Practicum Courses"
BRIEF = "India 2008/Project_Brief.txt"

NOTE_TEXT = "Removed uncertain issue area Education; see review."


def evidence(text, evidence_type="direct"):
    return {
        "source_file": BRIEF,
        "source_location": "page 1",
        "evidence_text": text,
        "evidence_type": evidence_type,
    }


def proposal(row, column_name, cell, value, confidence):
    return {
        "sheet": SHEET,
        "row": row,
        "column_name": column_name,
        "cell": cell,
        "value": value,
        "evidence": [evidence("Stated in the brief.")],
        "rules_applied": [],
        "confidence": confidence,
        "status": "proposed",
    }


def conflict_proposal(row, column_name, cell):
    return {
        "sheet": SHEET,
        "row": row,
        "column_name": column_name,
        "cell": cell,
        "value": None,
        "evidence": [evidence("Two source files disagree.")],
        "rules_applied": [],
        "confidence": None,
        "status": "conflict",
    }


FILLER_OUTPUT = {
    "proposals": [
        proposal(2, "Notes", "F2", "First draft note.", "high"),
        proposal(2, "Main Issue Area(s)", "G2", "Healthcare", "medium"),
        proposal(2, "Project ID*", "A2", "PRJ-0001", "medium"),
        proposal(2, "Start Date", "D2", "2026-01-01", "medium"),
        proposal(4, "Main Issue Area(s)", "G4", "Education", "medium"),
    ]
}


def finding(
    cell,
    verdict,
    recommended=None,
    comment="Reviewer comment.",
    missed_data=False,
):
    return {
        "cell": cell,
        "verdict": verdict,
        "recommended_value": recommended,
        "evidence": [evidence("Checked against the annual report.")],
        "reviewer_comment": comment,
        "missed_data": missed_data,
    }


PLANNED_CELLS = ("F2", "G2", "A2", "D2", "G4")


def complete_review(*findings):
    covered = {item["cell"] for item in findings}
    return {
        "findings": [
            *findings,
            *(finding(cell, "PASS") for cell in PLANNED_CELLS if cell not in covered),
        ]
    }


REVIEW_OUTPUT = {
    "findings": [
        finding("G2", "PASS"),
        finding("F2", "WARN", recommended="Corrected community note."),
        finding("A2", "WARN", recommended="PRJ-0002"),
        finding("G4", "FAIL", recommended="Healthcare"),
        finding("D2", "WARN", recommended="2026-02-02"),
        # Missed data: cell empty but determinable (FAIL -> FIX).
        finding("E2", "FAIL", recommended="Established", missed_data=True),
        # Reviewer could not adjudicate at all.
        finding(
            "D4",
            "UNRESOLVED",
            comment="Conflicting dates across sources.",
            missed_data=True,
        ),
    ]
}


def decision(cell, action, proposed=None, note_append=None, justification="Because."):
    return {
        "cell": cell,
        "action": action,
        "proposed_value": proposed,
        "note_append": note_append,
        "evidence": [evidence("Re-checked the original brief.")],
        "justification": justification,
    }


REVISION_OUTPUT = {
    "decisions": [
        decision("F2", "ACCEPT"),
        decision("A2", "REBUT", justification="Brief explicitly names PRJ-0001."),
        decision(
            "G4",
            "CLEAR",
            note_append=NOTE_TEXT,
            justification="Issue area cannot be determined from sources.",
        ),
        decision("D2", "REBUT", justification="Program page confirms the date."),
        # FIX: independent correction differing from the recommendation.
        decision("E2", "FIX", proposed="Emerging", justification="Brief page 3."),
        decision("D4", "UNRESOLVED", justification="Cannot adjudicate either."),
    ]
}

RE_REVIEW_OUTPUT = {
    "verdicts": [
        {"cell": "A2", "verdict": "WITHDRAWN", "reviewer_comment": "Convinced."},
        {"cell": "D2", "verdict": "UPHELD", "reviewer_comment": "Evidence stands."},
    ]
}


def start_run(inputs, review=None, runtimes=None):
    outputs = {
        "scoping": scoping_fixture(),
        "filler": FILLER_OUTPUT,
        "reviewer": review or REVIEW_OUTPUT,
        "revision": REVISION_OUTPUT,
        "re_review": RE_REVIEW_OUTPUT,
    }
    if runtimes is None:
        runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}
    return run_workflow(
        inputs=RunInputs(
            source=inputs["source"],
            workbook=inputs["workbook"],
            task=inputs["task"],
            rules_file=inputs["rules_file"],
            scoping_answers=inputs["scoping_answers"],
        ),
        runs_root=inputs["runs_root"],
        runtimes=runtimes,
    )


def workspace_of(state):
    return Path(state["workspace_path"])


def read_artifact(state, name):
    return json.loads((workspace_of(state) / "artifacts" / name).read_text())


def stage_names(state):
    with sqlite3.connect(workspace_of(state) / "state/audit.sqlite") as conn:
        rows = conn.execute(
            "SELECT stage FROM stages WHERE run_id = ? ORDER BY id",
            (state["run_id"],),
        ).fetchall()
    return [row[0] for row in rows]


def test_final_workbook_reflects_every_route(inputs):
    state = start_run(inputs)

    final = workspace_of(state) / "output/final.xlsx"
    assert Path(state["draft_xlsx_path"]).read_bytes() == final.read_bytes()
    sheet = load_workbook(final)[SHEET]

    assert sheet["F2"].value == "Corrected community note."  # ACCEPT
    assert sheet["G2"].value == "Healthcare"  # PASS frozen
    assert sheet["A2"].value == "PRJ-0001"  # REBUT withdrawn, original kept
    assert sheet["G4"].value is None  # CLEAR
    assert sheet["F4"].value == NOTE_TEXT  # note_append companion
    assert sheet["D2"].value == "2026-01-01"  # upheld rebuttal, no change
    assert sheet["E2"].value == "Emerging"  # FIX fills missed data
    assert sheet["D4"].value is None  # UNRESOLVED writes nothing
    assert sheet["H2"].value == "=SUM(1,2)"  # untouched formula


def test_pass_cells_receive_no_revision_mutations(inputs):
    state = start_run(inputs)

    with sqlite3.connect(workspace_of(state) / "state/audit.sqlite") as conn:
        rows = conn.execute(
            "SELECT cell FROM mutations WHERE actor_role = 'revision'"
            " AND status = 'applied'"
        ).fetchall()
    assert {row[0] for row in rows} == {"F2", "G4", "F4", "E2"}


def test_revision_receives_only_the_allowed_context(inputs):
    state = start_run(inputs)

    revision_inputs = json.loads(
        (workspace_of(state) / "agent_outputs/revision/inputs.json").read_text()
    )

    finding_cells = {item["cell"] for item in revision_inputs["findings"]}
    assert finding_cells == {"F2", "A2", "G4", "D2", "E2", "D4"}  # no PASS

    # Only cells that actually have a proposal/provenance carry one.
    assert set(revision_inputs["proposals"]) == {"F2", "A2", "G4", "D2"}
    assert set(revision_inputs["provenance"]) == {"F2", "A2", "G4", "D2"}
    assert revision_inputs["rules_dir"] == "input/rules"

    assert revision_inputs["mutation_allowlist"] == sorted(
        [
            f"{SHEET}!F2",
            f"{SHEET}!A2",
            f"{SHEET}!G4",
            f"{SHEET}!D2",
            f"{SHEET}!E2",
            f"{SHEET}!D4",
            f"{SHEET}!F4",  # Notes companion of flagged row 4
        ]
    )


def test_exactly_one_re_review_over_rebutted_cells_only(inputs):
    state = start_run(inputs)

    assert stage_names(state).count("CODEX_REREVIEW") == 1
    re_review = read_artifact(state, "re_review.json")
    assert {v["cell"] for v in re_review["verdicts"]} == {"A2", "D2"}


def test_unresolved_and_human_review_artifacts(inputs):
    state = start_run(inputs)

    unresolved = read_artifact(state, "unresolved.json")
    reasons = {item["cell"]: item["reason"] for item in unresolved["cells"]}
    assert set(reasons) == {"D2", "D4"}
    assert "upheld" in reasons["D2"]
    assert "revision" in reasons["D4"]

    human = read_artifact(state, "human_review.json")
    entries = {entry["cell"]: entry for entry in human["items"]}
    assert set(entries) == {"D2", "D4"}

    upheld = entries["D2"]
    assert upheld["current_value"] == "2026-01-01"
    assert upheld["reviewer"]["recommended_value"] == "2026-02-02"
    assert upheld["reviewer"]["comment"]
    assert upheld["revision"]["action"] == "REBUT"
    assert upheld["revision"]["justification"]
    assert upheld["re_review"]["verdict"] == "UPHELD"
    assert "upheld" in upheld["reason"]

    undetermined = entries["D4"]
    assert undetermined["current_value"] is None
    assert undetermined["revision"]["action"] == "UNRESOLVED"
    assert undetermined["re_review"] is None

    text = (workspace_of(state) / "artifacts/human_review.md").read_text()
    assert "D2" in text and "2026-02-02" in text
    # Both agents' evidence is readable without opening the JSON.
    assert "annual report" in text and "original brief" in text


def test_provenance_stays_in_sync_with_revised_cells(inputs):
    state = start_run(inputs)

    provenance = read_artifact(state, "provenance.json")
    entries = {entry["cell"]: entry for entry in provenance["entries"]}

    accept = entries[f"{SHEET}!F2"]
    assert accept["value"] == "Corrected community note."
    assert accept["agent_role"] == "revision"

    fixed = entries[f"{SHEET}!E2"]
    assert fixed["value"] == "Emerging"
    assert fixed["agent_role"] == "revision"

    cleared = entries[f"{SHEET}!G4"]
    assert cleared["value"] is None
    assert cleared["agent_role"] == "revision"

    note = entries[f"{SHEET}!F4"]
    assert note["value"] == NOTE_TEXT
    assert note["agent_role"] == "revision"

    untouched = entries[f"{SHEET}!A2"]
    assert untouched["value"] == "PRJ-0001"
    assert untouched["agent_role"] == "filler"


def test_review_artifacts_are_produced(inputs):
    state = start_run(inputs)

    review = read_artifact(state, "review.json")
    assert len(review["findings"]) == 7
    assert Path(state["review_path"]).name == "review.json"

    revision = read_artifact(state, "revision.json")
    assert len(revision["decisions"]) == 6

    for name in ("review.md", "revision_log.md"):
        assert (workspace_of(state) / "artifacts" / name).read_text()


def test_all_pass_review_short_circuits_to_finalize(inputs):
    all_pass = {"findings": [finding(cell, "PASS") for cell in PLANNED_CELLS]}
    # Only filler and reviewer runtimes exist: reaching revision or
    # re-review would raise KeyError inside the fake.
    runtimes = {
        "scoping": FakeAgentRuntime({"scoping": scoping_fixture()}),
        "filler": FakeAgentRuntime({"filler": FILLER_OUTPUT}),
        "reviewer": FakeAgentRuntime({"reviewer": all_pass}),
    }
    state = start_run(inputs, review=all_pass, runtimes=runtimes)

    stages = stage_names(state)
    for absent in (
        "CLAUDE_REVISE",
        "APPLY_ALLOWED_REVISIONS",
        "CODEX_REREVIEW",
        "HUMAN_REVIEW",
    ):
        assert absent not in stages

    final = load_workbook(workspace_of(state) / "output/final.xlsx")[SHEET]
    assert final["F2"].value == "First draft note."


def test_source_conflict_bypasses_revision_and_reaches_human_review(inputs):
    extraction = {"proposals": [conflict_proposal(4, "Start Date", "D4")]}
    review = {"findings": [finding("D4", "UNRESOLVED")]}
    runtimes = {
        "scoping": FakeAgentRuntime({"scoping": scoping_fixture()}),
        "filler": FakeAgentRuntime({"filler": extraction}),
        "reviewer": FakeAgentRuntime({"reviewer": review}),
    }

    state = start_run(inputs, runtimes=runtimes)

    stages = stage_names(state)
    assert "CLAUDE_REVISE" not in stages
    assert "APPLY_ALLOWED_REVISIONS" not in stages
    assert "HUMAN_REVIEW" in stages

    final = load_workbook(workspace_of(state) / "output/final.xlsx")[SHEET]
    assert final["D4"].value is None
    assert read_artifact(state, "unresolved.json") == {
        "cells": [
            {
                "cell": "D4",
                "reason": "protected source conflict requires human review",
            }
        ]
    }
    human = read_artifact(state, "human_review.json")
    assert human["items"][0]["cell"] == "D4"
    assert human["items"][0]["revision"] is None


def test_pass_finding_cannot_close_a_source_conflict(inputs):
    extraction = {"proposals": [conflict_proposal(4, "Start Date", "D4")]}
    review = {"findings": [finding("D4", "PASS")]}
    runtimes = {
        "scoping": FakeAgentRuntime({"scoping": scoping_fixture()}),
        "filler": FakeAgentRuntime({"filler": extraction}),
        "reviewer": FakeAgentRuntime({"reviewer": review}),
    }

    state = start_run(inputs, runtimes=runtimes)

    stages = stage_names(state)
    assert "CLAUDE_REVISE" not in stages
    assert "HUMAN_REVIEW" in stages
    assert read_artifact(state, "unresolved.json") == {
        "cells": [
            {
                "cell": "D4",
                "reason": "protected source conflict requires human review",
            }
        ]
    }
    (item,) = read_artifact(state, "human_review.json")["items"]
    assert item["reviewer"]["verdict"] == "PASS"
    assert item["revision"] is None


def test_revision_inputs_and_allowlist_exclude_source_conflicts(inputs):
    extraction = {
        "proposals": [
            proposal(4, "Main Issue Area(s)", "G4", "Education", "medium"),
            conflict_proposal(4, "Start Date", "D4"),
        ]
    }
    review = {
        "findings": [
            finding("G4", "FAIL", recommended="Healthcare"),
            finding("D4", "UNRESOLVED"),
        ]
    }
    revision = {
        "decisions": [
            decision(
                "G4",
                "CLEAR",
                note_append=NOTE_TEXT,
                justification="The issue area cannot be determined.",
            )
        ]
    }
    outputs = {
        "scoping": scoping_fixture(),
        "filler": extraction,
        "reviewer": review,
        "revision": revision,
    }
    runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}

    state = start_run(inputs, runtimes=runtimes)

    revision_inputs = json.loads(
        (workspace_of(state) / "agent_outputs/revision/inputs.json").read_text()
    )
    assert [item["cell"] for item in revision_inputs["findings"]] == ["G4"]
    assert set(revision_inputs["proposals"]) == {"G4"}
    assert all("D4" not in item for item in revision_inputs["mutation_allowlist"])

    unresolved = read_artifact(state, "unresolved.json")
    assert unresolved["cells"] == [
        {
            "cell": "D4",
            "reason": "protected source conflict requires human review",
        }
    ]


def test_revision_decision_for_a_protected_source_conflict_is_illegal(inputs):
    extraction = {
        "proposals": [
            proposal(4, "Main Issue Area(s)", "G4", "Education", "medium"),
            conflict_proposal(4, "Start Date", "D4"),
        ]
    }
    review = {
        "findings": [
            finding("G4", "FAIL", recommended="Healthcare"),
            finding("D4", "UNRESOLVED"),
        ]
    }
    illegal_revision = {
        "decisions": [decision("D4", "CLEAR", note_append="Tried to clear a conflict.")]
    }
    outputs = {
        "scoping": scoping_fixture(),
        "filler": extraction,
        "reviewer": review,
        "revision": illegal_revision,
    }
    runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}

    with pytest.raises(ValueError, match="no matching finding"):
        start_run(inputs, runtimes=runtimes)


def test_same_batch_note_appends_share_one_notes_cell(inputs):
    # Two non-PASS findings on the same row, both decisions carrying a
    # note_append: the notes compose in decision order on the shared
    # Notes cell instead of clobbering each other.
    second_note = "Start date corrected against the program page."
    review = complete_review(
        finding("G4", "FAIL", recommended="Healthcare"),
        finding("D4", "FAIL", recommended="2026-04-04", missed_data=True),
    )
    revision = {
        "decisions": [
            decision(
                "G4",
                "CLEAR",
                note_append=NOTE_TEXT,
                justification="Issue area cannot be determined from sources.",
            ),
            decision(
                "D4",
                "FIX",
                proposed="2026-04-04",
                note_append=second_note,
                justification="Program page states the date.",
            ),
        ]
    }
    outputs = {
        "scoping": scoping_fixture(),
        "filler": FILLER_OUTPUT,
        "reviewer": review,
        "revision": revision,
    }
    runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}
    state = start_run(inputs, runtimes=runtimes)

    sheet = load_workbook(workspace_of(state) / "output/final.xlsx")[SHEET]
    assert sheet["G4"].value is None
    assert sheet["D4"].value == "2026-04-04"
    assert sheet["F4"].value == f"{NOTE_TEXT}\n{second_note}"

    # The audit trail records both applied writes, each old value being
    # the previous composed state of the shared cell.
    with sqlite3.connect(workspace_of(state) / "state/audit.sqlite") as conn:
        rows = conn.execute(
            "SELECT old_value, new_value FROM mutations"
            " WHERE cell = 'F4' AND status = 'applied' ORDER BY id",
        ).fetchall()
    assert [(json.loads(old), json.loads(new)) for old, new in rows] == [
        (None, NOTE_TEXT),
        (NOTE_TEXT, f"{NOTE_TEXT}\n{second_note}"),
    ]


def test_clear_on_the_notes_cell_itself_completes_the_run(inputs):
    # The revision prompt REQUIRES a note_append on every CLEAR — so a
    # FAIL finding on a Notes cell steers straight into the
    # self-targeting shape. The batch must apply, not abort.
    review = complete_review(finding("F2", "FAIL"))
    revision = {
        "decisions": [
            decision(
                "F2",
                "CLEAR",
                note_append=NOTE_TEXT,
                justification="Unsupported note; context preserved here.",
            )
        ]
    }
    outputs = {
        "scoping": scoping_fixture(),
        "filler": FILLER_OUTPUT,
        "reviewer": review,
        "revision": revision,
    }
    runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}
    state = start_run(inputs, runtimes=runtimes)

    sheet = load_workbook(workspace_of(state) / "output/final.xlsx")[SHEET]
    assert sheet["F2"].value == NOTE_TEXT

    with sqlite3.connect(workspace_of(state) / "state/audit.sqlite") as conn:
        rows = conn.execute(
            "SELECT new_value FROM mutations WHERE cell = 'F2'"
            " AND actor_role = 'revision' AND status = 'applied'",
        ).fetchall()
    assert [json.loads(value) for (value,) in rows] == [NOTE_TEXT]


def test_illegal_decision_fails_the_run(inputs):
    bad_revision = {
        "decisions": [
            decision("F2", "FIX", proposed="sneaky fix")
        ]  # WARN needs ACCEPT/REBUT
    }
    outputs = {
        "scoping": scoping_fixture(),
        "filler": FILLER_OUTPUT,
        "reviewer": REVIEW_OUTPUT,
        "revision": bad_revision,
        "re_review": RE_REVIEW_OUTPUT,
    }
    runtimes = {role: FakeAgentRuntime(outputs) for role in outputs}
    with pytest.raises(ValueError, match="WARN"):
        start_run(inputs, runtimes=runtimes)
